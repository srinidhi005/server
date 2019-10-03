import json

import openpyxl
import requests
import sys

def write_to_excel(data):
    """
    Writes to excel
    :param statement_data:
    :return:
    """
    template_file = sys.argv[1] + "/1ExtractionGSheet_Template.xlsx"
    company = "Nike"
    wb = openpyxl.load_workbook(filename=template_file)
    ws = wb.worksheets[0]
    col_ref = 3
    for key in data.keys():
        ws.cell(row=1, column=col_ref).value = key
        ws.cell(row=3, column=col_ref).value = data[key]["AM_IS_I"]
        ws.cell(row=7, column=col_ref).value = data[key]["AM_IS_GP"]
        ws.cell(row=10, column=col_ref).value = data[key]["AM_IS_EXP"]
        ws.cell(row=11, column=col_ref).value = data[key]["EB2"]
        ws.cell(row=13, column=col_ref).value = data[key]["AM_IS_DEP_AMO"]
        ws.cell(row=14, column=col_ref).value = data[key]["EB1"]
        ws.cell(row=17, column=col_ref).value = data[key]["EB2"]
        ws.cell(row=18, column=col_ref).value = data[key]["AM_IS_NIEXP"]
        ws.cell(row=19, column=col_ref).value = data[key]["AM_IS_OE"]
        ws.cell(row=20, column=col_ref).value = data[key]["EB3"]
        ws.cell(row=21, column=col_ref).value = data[key]["AM_IS_TX"]
        ws.cell(row=22, column=col_ref).value = data[key]["AM_IS_NI"]
        col_ref += 1
        if col_ref > 4:break
    wb.save(sys.argv[1] + "/" + company + ".xlsx")

def read_json(url):
    response = requests.get(url)
    response_json = json.loads(response.text)
    return response_json

json_url = "https://storage.googleapis.com/extraction-engine/2ExtractionJSON/file0.json"
response_json = read_json(json_url)
# print(response_json["period"])
company_data = {}
for data in response_json["period"]:
    # print(data["asof"])
    statistics = {}
    # print(data["Additional"])
    for code in data["Additional"]:
        if code["code"] == "EB1":
            EB1=code["value"]
        if code["code"] == "EB2":
            EB2=code["value"]
        if code["code"] == "EB3":
            EB3=code["value"]
        if code["code"] == "AM_IS_EXP":
            AM_IS_EXP=code["value"]
        if code["code"] == "AM_IS_DEP_AMO":
            AM_IS_DEP_AMO=code["value"]
        if code["code"] == "AM_IS_NI":
            AM_IS_NI=code["value"]
        if code["code"] == "AM_IS_TX":
            AM_IS_TX=code["value"]
        if code["code"] == "AM_IS_NIEXP":
            AM_IS_NIEXP=code["value"]
        if code["code"] == "AM_IS_OE":
            AM_IS_OE=code["value"]
    for code in data["statement"]:
        if code["code"] == "AM_IS_I":
            AM_IS_I = code["value"]
        if code["code"] == "AM_IS_GP":
            AM_IS_GP = code["value"]
    statistics["AM_IS_I"] = AM_IS_I
    statistics["AM_IS_GP"] = AM_IS_GP
    statistics["AM_IS_DEP_AMO"] = AM_IS_DEP_AMO
    statistics["AM_IS_EXP"] = AM_IS_EXP
    statistics["AM_IS_NI"] = AM_IS_NI
    statistics["AM_IS_OE"] = AM_IS_OE
    statistics["AM_IS_TX"] = AM_IS_TX
    statistics["AM_IS_NIEXP"] = AM_IS_NIEXP
    statistics["EB1"] = EB1
    statistics["EB2"] = EB2
    statistics["EB3"] = EB3
    company_data[data["asof"]] = statistics
# print(company_data)


write_to_excel(company_data)

# https://storage.googleapis.com/extraction-engine/2ExtractionJSON/file0.json
