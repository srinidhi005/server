# -*- coding: utf-8 -*-

dicti = {'Cash Equivalents': ['Cash', 'Cash Equivalents', 'Cash & Equivalents', 'Cash and cash equivalents',
                              'Cash and equivalents', 'Cash & Cash Equivalents', 'Money Market',
                              'Money Market Securities', 'Marketable securities'],
         'Shareholders Equity': ['Shareholders’ Equity', 'Stockholders’ Equity', 'Owners’ Equity', 'Equity'],
         'Net Income/Loss': ['Net loss (Income)', 'Net loss/Income', 'Net Income (loss)', 'Net Income/loss',
                             'Net Income', 'Net loss'],
         'Retained Earnings': ['Retained earnings', 'Accumulated Deficit', 'Retained earnings (Accumulated Deficit)',
                               'Accumulated Deficit (Retained earnings)'],
         'Accounts Receivable': ['Receivables', 'Accounts receivable', 'Accounts receivable,net', 'AR'],
         'Inventories': ['Inventories', 'Inventories - Total', 'Inventory'],
         'Other current assets': ['Prepaid Expenditures', 'Prepayments', 'Prepayables', 'Prepaid Expenses',
                                  'Prepaid Expenses and other current assets', 'Current Assets-Other',
                                  'Other Current Assets', 'Other Short Term Current Assets',
                                  'Other Short-Term Current Assets'],
         'Total Current Assets': ['Current Assets-Total', 'Total Current Assets'],
         'PPE': ['Plant,Property & Equip(Gross)', 'Plant,Property & Equipment', 'Plant,Property and Equipment', 'PPE'],
         'PPE(Net)': ['Property,plant and equipment,net', 'Plant,Property & Equipment,net',
                      'Property,plant and equipment,less accumulated depreciation', 'Property and equipment,net',
                      'Plant and equipment,net', 'Plant,Property & Equip(net)', 'Plant,property & Equipment(Net)',
                      'Plant,property and Equipment(Net)', 'PPE(Net)'],
         'Intangible Assets': ['Identifiable intangible assets, net', 'Other intangible assets',
                               'Intangible assets, net',
                               'Identifiable intangible assets, less accumulated amortization',
                               'Other intangible assets, net'],
         'Net Revenue': ['Net Sales', 'Net Revenue', 'Net Revenues', 'Revenue', 'Revenues', 'Net Sale'],
         'Cost of Revenue': ['Cost of Sales', 'Cost of Revenues', 'Cost of Sale', 'Cost of Revenue', 'COGs',
                             'Cost of goods sold'],
         'Gross Profit': ['Gross Margin', 'Gross Profit'],
         'Operating Income/loss': ['Operating income', 'Operating loss', 'Operating income (loss)', 'Operating Profit',
                                   'Income from operations', 'Operating Income/Loss', 'Loss from operations',
                                   'Profit from operations', 'Operating income or loss'],
         'Total Operating Expenses': ['Total operating expenses', 'Operating expenses'],
         'Research and Development': ['Research and development', 'R & D'],
         'Earnings before Income Taxes': ['Loss before income taxes', 'Income before provision for income taxes',
                                          'Earnings before income taxes', 'income (loss) before income taxes'],
         'Provision for income taxes': ['Income tax expense', 'Income tax expense (benefit)',
                                        'provison/(benefit) for taxes on income', 'Income tax benefit (provision)',
                                        'Provision for income taxes'],
         'Net change in cash': ['Net increase (decrease) in cash and cash equivalents',
                                'Net increase (decrease) in cash, cash equivalents, and restricted cash',
                                'net decrease (increase) in cash and cash equivalents',
                                'Increase/(Decrease) in cash and cash equivalents',
                                'Net increase/(Decrease) in cash and cash equivalents',
                                'Net increase (decrease) in cash and equivalents',
                                'Cash and cash equivalents: Increase', 'Net decrease in cash and cash equivalents',
                                'Net increase in cash and cash equivalents'],
         'Cash Beginning': ['Cash and cash equivalents at beginning of period',
                            'Cash, cash equivalents, and restricted cash at beginning of period',
                            'Cash and cash equivalents-beginning of year',
                            'Cash and cash equivalents , beginning of period',
                            'Cash and equivalents, beginning of year',
                            'Cash and cash equivalents: Balance at beginning of year',
                            'Cash and cash equivalents, beginning', 'Cash and cash equivalents at beginning of year'],
         'Cash Ending': ['Cash and cash equivalents at end of period',
                         'Cash, cash equivalents, and restricted cash at end of period',
                         'Cash and cash equivalents, end', 'Cash and cash equivalents , end of period',
                         'Cash and equivalents, end of year', 'Cash and cash equivalents-end of year',
                         'Cash and cash equivalents: Balance at end of year',
                         'Cash and cash equivalents at end of year', 'Cash and cash equivalents , end of the year']
         }

code_dict = {'Net Revenue': 'AM_IS_I',
             'Gross Profit': 'AM_IS_GP',
             'Operating Income/loss': 'AM_IS_NOI',
             'Total Operating Expenses': 'AM_IS_E',
             'Cost of Revenue': 'AM_IS_CORS',
             'Net Income/Loss': 'AM_IS_NI',
             'Cash Equivalents': 'AM_BS_BA',
             'Shareholders Equity': 'AM_BS_E',
             'Net Income/Loss': 'AM_BS_NI',
             'Accounts Receivable': 'AM_BS_AR',
             'Other current assets': 'AM_BS_OCA',
             'Total Current Assets': 'AM_BS_CA',
             'Net change in cash': 'AM_CF_CI',
             'Cash Beginning': 'AM_CF_BC',
             'Cash Ending': 'AM_CF_EC'
             }

#########################################################################


from spellchecker import SpellChecker
import re
import sys
import json

from test import parseet
from db_connection import *
from datetime import datetime

import nltk

import pandas as pd
import tabula
import os
import PyPDF2
import itertools

import decimal
import openpyxl
import requests

cwd_path = os.getcwd()
CWD_PATH = os.getcwd()

file = sys.argv[1]
# file_name1=sys.argv[2]


gross = []
total_expense = []
inc_tax = []
inc_expense = []
ebitda = []
ebit = []
ebt = []
net_inc_loss = []
dp = []
other_expense = []
k1 = 0

months_dict = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october',
               'november', 'december']
list_folder = [sys.argv[3] + '/text', sys.argv[3] + '/json', sys.argv[3] + '/firsttext']
for i in list_folder:
    for th in os.listdir(os.path.join(cwd_path, i)):
        file_path_del = os.path.join(cwd_path, i, th)
        try:
            if os.path.isfile(file_path_del):
                os.unlink(file_path_del)
        except Exception as e:
            print(e)

# path = os.path.join(cwd_path,'pdf',file)
path = file

from PyPDF2 import PdfFileReader

pdf = PdfFileReader(open(path, 'rb'))
pages = pdf.getNumPages()

import pdftotext

with open(file, 'rb') as f:
    pdf = pdftotext.PDF(f, raw=False)

for p in pdf:
    with open(sys.argv[3] + '/firsttext/ne.txt', 'a') as f:
        f.write(p)

file_name1 = sys.argv[3] + '/firsttext/ne.txt'


###########################

def fo(li):
    line = li.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    return line


#####################################################################################

title, company, units, currency, typee, dates = parseet(file_name1)
numstat = len(typee)
cul = []
for i in dates:
    if not isinstance(i[0], str) and i[0] > 31:
        cul.append(len(i))

num = 0
blck = []
num1 = 0
nt = dates
for m, k in enumerate(dates):
    if cul:
        if isinstance(k[0], str):
            num1 += 1
            if len(k) < cul[num]:
                v = int(cul[num]) - len(k)
                for c in range(v):
                    nt[m].append(k[0])
        elif k[0] < 32:
            num1 += 1
            if len(k) < cul[num]:
                v = int(cul[num]) - len(k)
                for c in range(v):
                    nt[m].append(k[0])
        else:
            num += 1
            num1 += 1
            blck.append(num1)
            num1 = 0

    else:
        break


###########################################################################################################

def sub_items(filepath):
    indentation = []
    indentation.append(0)
    depth = 0
    depth_lis = []
    content_lis = []

    with open(filepath, encoding="utf8") as infile:
        for line in infile:
            line = line[:-1]
            c = re.split(r'\s{2,}', line.strip())

            if c:
                content_lis.append(c)

            content = line.strip()
            indent = len(line) - len(content)

            if indent > indentation[-1]:
                depth += 1
                indentation.append(indent)

            elif indent < indentation[-1]:
                while indent < indentation[-1]:
                    depth -= 1
                    indentation.pop()

                if indent != indentation[-1]:
                    pass
                    # raise RuntimeError("Bad formatting")
            depth_lis.append(depth)
    return depth_lis, content_lis


####################################################################################

type_len = len(typee)
main_json = {}
json_collection = []
out_m = 0

start_line = 0

note_lis = []
unit_convers = False
type_var = False
months_var = False
#################################################################    1   ################################################

for i in range(pages):
    df = tabula.read_pdf(path, pages=i + 1, encoding='cp1252', multiple_table=True)
    if df is not None and isinstance(df, pd.DataFrame) and not df.empty:
        spell = SpellChecker()
        newlist = []
        codelist = []
        suggestion_flag = []
        for text in df.iloc[:, 0]:

            flag = 0

            text = re.sub('[^a-zA-Z0-9\n\.]', ' ', str(text))
            text = re.sub("[\(\[].*?[\)\]]", "", text)
            text = re.sub(' +', " ", text)
            if re.search('accounts receivable', text, re.I):
                text = "Accounts Receivable"
            text_vocab = list(w.lower() for w in text.split() if w.isalpha())

            translated = ""
            code = ""
            flag = 0
            temp = []
            for word in text_vocab:

                temp.append(word)
                misspelled = spell.unknown(temp)
                temp = []
                if misspelled:
                    flag = 1
                    translated = translated + " " + word
                else:
                    translated = translated + " " + word
            translated = translated.strip()

            for key, value in dicti.items():
                for item in value:

                    item = re.sub('[^a-zA-Z0-9\n\.]', ' ', item)
                    item = re.sub("[\(\[].*?[\)\]]", "", item)
                    item = re.sub(' +', " ", item)
                    if translated == item.lower():
                        translated = key
                        code = code_dict.get(translated, "None")

            translated = translated.capitalize()
            codelist.append(code)
            newlist.append(translated)
            suggestion_flag.append(flag)
        tdf = pd.DataFrame()
        tdf['Original'] = df.iloc[:, 0]
        tdf['Transformed'] = newlist
        tdf['Code'] = codelist
        tdf['suggestion'] = suggestion_flag
        tdf = tdf.dropna()

        ###########################################################################  2 #######################################################

        dataset = tdf
        dataset = dataset.fillna('None')
        dataset = dataset[~dataset.Original.str.replace('(', ' ').str.contains('Note')]

        checkt1 = str(dataset.iloc[0, 0]).replace('(', ' ').replace(')', ' ').replace(',', ' ')

        for months_i in months_dict:
            if re.search('\\b' + months_i + '\\b', checkt1, re.I):
                months_var = True
            else:
                months_var = False

        if re.search('\\b' + 'millions' + '\\b', checkt1, re.I) or re.search('\\b' + 'thousands' + '\\b', checkt1,
                                                                             re.I) or re.search('\\b' + 'per' + '\\b',
                                                                                                checkt1,
                                                                                                re.I) or re.search(
                '\\b' + 'par' + '\\b', checkt1, re.I) or months_var:
            dataset = dataset.drop(dataset.index[0])

        original1 = list(dataset.iloc[:, 0])
        transformed = list(dataset.iloc[:, 1])
        code = list(dataset.iloc[:, 2])
        col = list(dataset.iloc[:, 0])

        for en, va in enumerate(code):
            if va == "None":
                code[en] = ''

        with open(file_name1, encoding="utf8") as infile, open(sys.argv[3] + "/text/fileer.txt".format(out_m), "w",
                                                               encoding="utf8") as outfile:
            for n, line in enumerate(infile):
                if col:

                    line1 = re.sub('[^a-zA-Z0-9\n\.]', ' ', str(line))
                    line1 = re.sub("[\(\[].*?[\)\]]", "", line1)
                    line1 = re.sub(' +', " ", line1)

                    text = re.sub('[^a-zA-Z0-9\n\.]', ' ', col[0])
                    text = re.sub("[\(\[].*?[\)\]]", "", text)
                    text = re.sub(' +', " ", text)
                    check1 = text.replace('(', " ")

                    if re.search(text, line1, re.I) and n >= start_line:
                        outfile.write(fo(line))
                        col.pop(0)
                else:
                    start_line = n
                    break

        #########################################################################         3         #######################################

        folder1 = sys.argv[3] + '/text'
        the_file = 'fileer.txt'
        n = 0
        file_path = os.path.join(CWD_PATH, folder1, the_file)
        indentation = []
        indentation.append(0)
        depth = 0
        lis = []
        main_json['title'] = title[out_m]
        main_json["company"] = company
        if units:
            main_json["units"] = units.replace('(', '').replace(')', '')
        else:
            main_json["units"] = ''

        main_json["currency"] = currency
        main_json["type"] = typee[out_m]
        if main_json["type"] == "Balance_sheet":
            type_var = True
            inc_flag = False
            cf_flag = False
        if main_json["type"] == "statement_of_income":
            inc_flag = True
            cf_flag = False
            flag_set = False
        if main_json["type"] == "cash_flow_statement":
            inc_flag = False
            cf_flag = True
            flag_set = False
        main_json['period'] = [{} for i in range(cul[out_m])]

        if main_json["units"] == 'millions':
            unit_convers = True

        if blck:
            n = blck[0]
            blck.pop(0)

        for k, j in enumerate(main_json['period']):
            st = ""
            flagg = 0

            for z in range(n):
                st += str(nt[z][k])
            j["asof"] = st
            j['statement'] = []
            depth_lis, content_lis = sub_items(file_path)

            conversio = False
            line_span1 = False
            sum_inc = 0
            sum_dp = 0

            for f, b, ori, tr, co in zip(depth_lis, content_lis, original1, transformed, code):

                if unit_convers:
                    if re.search('per common share', b[0], re.I) or re.search('per share', b[0], re.I) or re.search(
                            'par value', b[0], re.I):
                        conversio = True
                        if len(b) == 1:
                            line_span1 = True
                        else:
                            line_span1 = False
                    elif line_span1:
                        if len(b) == 1:
                            conversio = False
                            line_span1 = False
                    else:
                        conversio = False

                if inc_flag == True and len(b) != 1:
                    '''
                    if re.search('gross',b[0],re.I) and re.search('profit',b[0],re.I):
                        gross.insert(k,b[k+1])
                    elif re.search('expenses',b[0],re.I) or re.search('expense',b[0],re.I) and re.search('total',b[0],re.I):
                        total_expense.insert(k,b[k+1])
                    '''
                    if re.search('other', tr, re.I) and re.search('expense', tr, re.I):
                        other_expense.insert(k, b[k + 1])
                        flagg = 1
                    if re.search('gross profit', tr, re.I) or re.search('Operating Income/loss', tr, re.I):
                        gross.insert(k, b[k + 1])
                    if flagg == 0:
                        if (re.search('income', b[0], re.I) or re.search('interest', b[0], re.I)) and (
                                re.search('expense', b[0], re.I) or re.search('expenses', b[0], re.I)) and not (
                                re.search('taxes', b[0], re.I) or re.search('tax', b[0], re.I)):

                            sum_inc += float(b[k + 1])
                            if len(inc_expense) > k:
                                inc_expense[k] = sum_inc

                            else:
                                inc_expense.insert(k, sum_inc)

                    elif re.search('income', b[0], re.I) and (
                            re.search('tax', b[0], re.I) or re.search('taxes', b[0], re.I)) and (
                            re.search('expense', b[0], re.I) or re.search('expenses', b[0], re.I)) or re.search(
                            "Provision for income taxes", tr, re.I):

                        inc_tax.insert(k, b[k + 1])
                    elif re.search('^net income$', b[0], re.I) or re.search('^net income/loss$', b[0],
                                                                            re.I) or re.search('^net loss$', b[0],
                                                                                               re.I) or re.search(
                            '^net loss/income$', b[0], re.I) or re.search('^net loss (income)$', b[0],
                                                                          re.I) or re.search('^net income (loss)$',
                                                                                             b[0], re.I):

                        net_inc_loss.insert(k, b[k + 1])
                    if re.search('depreciation', b[0], re.I) or re.search('amortization', b[0], re.I):
                        flag_set = "True"
                        sum_dp += float(b[k + 1])
                        if len(dp) > k:
                            dp[k] = sum_dp

                if cf_flag == True and flag_set == False:
                    if re.search('depreciation', b[0], re.I) or re.search('amortization', b[0], re.I):
                        sum_dp += float(b[k + 1])
                        if len(dp) > k:
                            dp[k] = sum_dp

                        else:
                            dp.insert(k, sum_dp)

                if re.search('Total', b[0], re.I) and len(b) > 1 and type_var:
                    if k + 1 < len(b):
                        b[k + 1] = re.sub('[^A-Za-z0-9.-]+', '', b[k + 1])
                        try:

                            b[k + 1] = float(b[k + 1])
                            if unit_convers and not conversio:
                                b[k + 1] = b[k + 1] * 1000

                        except:
                            pass

                        total_c = {"desc": b[0], 'transformed': tr, 'code': co, 'value': b[k + 1]}
                        v = j['statement'][-1]
                        if 'total' in v:
                            v['total1'] = total_c
                        else:

                            v['total'] = total_c
                else:

                    if len(b) == 1:

                        head_d = {'heading': b[0], 'transformed': tr, 'code': co, 'items': []}
                        if head_d['heading'] or head_d['items']:
                            j['statement'].append(head_d)
                    elif f != 0:
                        if k + 1 < len(b):
                            b[k + 1] = re.sub('[^A-Za-z0-9.-]+', '', b[k + 1])
                            try:

                                b[k + 1] = float(b[k + 1])
                                if unit_convers and not conversio:
                                    b[k + 1] = b[k + 1] * 1000

                            except:
                                pass

                            decs_c = {"desc": b[0], 'transformed': tr, 'code': co, 'value': b[k + 1]}
                            if not j['statement']:
                                j['statement'].append(decs_c)
                            else:
                                desc_v = j['statement'][-1]
                                if 'items' in desc_v:
                                    desc_v['items'].append(decs_c)
                                else:
                                    j['statement'].append(decs_c)

                    else:
                        if k + 1 < len(b):
                            b[k + 1] = re.sub('[^A-Za-z0-9.-]+', '', b[k + 1])
                            try:

                                b[k + 1] = float(b[k + 1])
                                if unit_convers and not conversio:
                                    b[k + 1] = b[k + 1] * 1000


                            except:
                                pass

                            desc_l = {"desc": b[0], 'transformed': tr, 'code': co, 'value': b[k + 1]}
                            j['statement'].append(desc_l)
            k1 = k

        del nt[:n]
        #with open(sys.argv[2] + "{}.json".format(out_m), "w") as outfile:
        #    json.dump(main_json, outfile, indent=4)
        json_collection.append(main_json)
        out_m += 1
        main_json = {}

#######################################################################################
# print(inc_tax,inc_expense,net_inc_loss,dp)
if not net_inc_loss:
    net_inc_loss = [0] * (k1 + 1)
if not other_expense:
    other_expense = [0] * (k1 + 1)
if not dp:
    dp = [0] * (k1 + 1)
if not inc_expense:
    inc_expense = [0] * (k1 + 1)
if not inc_tax:
    inc_tax = [0] * (k1 + 1)
for m, n, p, i, g, o in zip(dp, inc_expense, inc_tax, net_inc_loss, gross, other_expense):
    if unit_convers:

        val = (float(i) + float(p)) * 1000
        ebt.append(val)
        val2 = val + float(n) * 1000
        ebit.append(val2)

        val3 = val2 + float(m) * 1000
        ebitda.append(val3)

        val4 = float(g) * 1000 - val3
        total_expense.append(val4)






    else:
        val = float(i) + float(p)
        ebt.append(val)
        val2 = val + float(n)
        ebit.append(val2)

        val3 = val2 + float(m)
        ebitda.append(val3)

        val4 = float(g) - val3
        total_expense.append(val4)

if unit_convers:
    dp = [i * 1000 for i in dp]
    inc_expense = [i * 1000 for i in inc_expense]
    net_inc_loss = list(map(lambda x: float(x) * 1000, net_inc_loss))
    inc_tax = list(map(lambda x: float(x) * 1000, inc_tax))
    other_expense = list(map(lambda x: float(x) * 1000, other_expense))

# if ebitda and ebit and ebt:
folder = sys.argv[2].replace('/file', '')
folder = folder.replace('./', '')


# Write to Excel file

def write_to_excel(data):
    """
    Writes to excel
    :param statement_data:
    :return:
    """
    template_file = sys.argv[3] + "/1ExtractionGSheet_Template.xlsx"
    company = "Nike"
    wb = openpyxl.load_workbook(filename=template_file)
    ws = wb.worksheets[0]
    col_ref = 3
    for key in data.keys():
        ws.cell(row=1, column=col_ref).value = key
        ws.cell(row=10, column=col_ref).value = data[key]["AM_IS_EXP"]
        ws.cell(row=11, column=col_ref).value = data[key]["EB2"]
        ws.cell(row=13, column=col_ref).value = data[key]["AM_IS_DEP_AMO"]
        ws.cell(row=14, column=col_ref).value = data[key]["EB1"]
        ws.cell(row=17, column=col_ref).value = data[key]["EB2"]
        ws.cell(row=18, column=col_ref).value = data[key]["AM_IS_NIEXP"]
        ws.cell(row=19, column=col_ref).value = data[key]["AM_IS_OE"]
        ws.cell(row=20, column=col_ref).value = data[key]["EB3"]
        ws.cell(row=21, column=col_ref).value = data[key]["AM_IS_TX"]
        ws.cell(row=22, column=col_ref).value = data[key]["AM_IS_NI"]
        col_ref += 1

    wb.save(sys.argv[3] + "/" + company + ".xlsx")


for d in json_collection:
    if d['type'] == "statement_of_income":
        for i,j,k,a,b,c,d1,m,t,o in zip(ebitda,ebit,ebt,net_inc_loss,dp,inc_expense,inc_tax,d['period'],total_expense,other_expense):
            m['Additional']=[{"desc":"(-) Selling, General, & Administrative ('SG&A')","code":"AM_IS_EXP","value":t},{"desc":"EBITDA","code":"EB1","value":i},{"desc":"EBIT","code":"EB2","value":j},{"desc":"(+) Depreciation & Amortization ('D&A')","code":"AM_IS_DEP_AMO","value":b},{"desc":"EBT","code":"EB3","value":k},{"desc":"(-) Net Interest Expense","code":"AM_IS_NIEXP","value":c},{"desc":"Net Income","code":"AM_IS_NI","value":float(a)},{"desc":"(+ / -) Other Income / (Expense)","code":"AM_IS_OE","value":o},{"desc":"(-) Taxes","code":"AM_IS_TX","value":float(d1)}]
        json_copy = d
        print("*" * 30)
        print(d)
        print("*" * 30)


##########################################################


def inject_db(json_data,latest_enum):
    con = db_connect()  # connect to database
    if con is not None:
        cursor = con.cursor()
        query = "delete from company_actuals where companyname='"+sys.argv[4]+"'"
        cursor.execute(query)
        con.commit()
        for data in json_data["period"]:
            if data["asof"] ==  latest_enum[0] :
                continue
            for key,val in latest_enum.items():
                # print(key, val)
                if val.lower() == (data["asof"]).lower():
                    latest = key
                    break

            for code in data["Additional"]:
                if code["code"] == "AM_IS_EXP":
                    AM_IS_EXP = int(code["value"])
                if code["code"] == "AM_IS_DEP_AMO":
                    AM_IS_DEP_AMO = int(code["value"])
                if code["code"] == "AM_IS_TX":
                    AM_IS_TX = int(code["value"])
                if code["code"] == "AM_IS_NIEXP":
                    AM_IS_NIEXP = int(code["value"])
                if code["code"] == "AM_IS_OE":
                    AM_IS_OE = int(code["value"])

            for code in data["statement"]:
                if code["code"] == "AM_IS_I":
                    AM_IS_I = int(code["value"])
                if code["code"] == "AM_IS_CORS":
                    AM_IS_CORS = int(code["value"])
            gross_profit = AM_IS_I - AM_IS_CORS
            ebit = gross_profit - AM_IS_EXP
            ebitda = ebit + AM_IS_DEP_AMO
            ebt = ebit - AM_IS_NIEXP - AM_IS_OE
            netincome = ebt - AM_IS_TX
            grossprofitmargin = (gross_profit / AM_IS_I) * 100
            ebitmargin = (ebit / AM_IS_I) * 100
            ebitdamargin = (ebitda / AM_IS_I) * 100
            ebtmargin = (ebt / AM_IS_I) * 100
            netincomemargin = (netincome / AM_IS_I) * 100

            query = "insert into company_actuals (companyname,asof,latest,totalrevenue,cogs,sga,da,netinterest,otherincome," \
                    "taxes,grossprofit,ebit,ebitda,netincome,grossprofitmargin,ebitmargin,ebitdamargin,ebtmargin,netincomemargin) values(" \
                    "'" + sys.argv[4] + "'," +str(int(str(datetime.strptime(data["asof"], "%b%d%Y"))[:4])) + "," + str(
                latest) + "," + str(AM_IS_I) + "," + str(AM_IS_CORS) + "," + str(AM_IS_EXP) + "," + str(
                AM_IS_DEP_AMO) + "," + str(AM_IS_NIEXP) + "," + str(AM_IS_OE) + "," + str(AM_IS_TX) + "," + str(gross_profit) + "," + str(ebit) + "" \
              "," + str(ebitda) + "," + str(netincome) + "," + str(grossprofitmargin) + "," + str(ebitmargin) + "," + str(ebitdamargin) + "" \
              "," + str(ebitmargin) + "," + str(netincomemargin) + ")"
            cursor.execute(query)
            con.commit()
        for data in json_data["period"]:
            if data["asof"] ==  latest_enum[0] :
                for key, val in latest_enum.items():
                    # print(key,val)
                    if val.lower() == (data["asof"]).lower():
                        latest = key
                        break

                for code in data["Additional"]:
                    if code["code"] == "AM_IS_EXP":
                        AM_IS_EXP = int(code["value"])
                    if code["code"] == "AM_IS_DEP_AMO":
                        AM_IS_DEP_AMO = int(code["value"])
                    if code["code"] == "AM_IS_TX":
                        AM_IS_TX = int(code["value"])
                    if code["code"] == "AM_IS_NIEXP":
                        AM_IS_NIEXP = int(code["value"])
                    if code["code"] == "AM_IS_OE":
                        AM_IS_OE = int(code["value"])

                for code in data["statement"]:
                    if code["code"] == "AM_IS_I":
                        AM_IS_I = int(code["value"])
                    if code["code"] == "AM_IS_CORS":
                        AM_IS_CORS = int(code["value"])
                gross_profit = AM_IS_I - AM_IS_CORS
                ebit = gross_profit - AM_IS_EXP
                ebitda = ebit + AM_IS_DEP_AMO
                ebt = ebit - AM_IS_NIEXP - AM_IS_OE
                netincome = ebt - AM_IS_TX
                grossprofitmargin = (gross_profit / AM_IS_I) * 100
                ebitmargin = (ebit / AM_IS_I) * 100
                ebitdamargin = (ebitda / AM_IS_I) * 100
                ebtmargin = (ebt / AM_IS_I) * 100
                netincomemargin = (netincome / AM_IS_I) * 100

                query = "insert into company_actuals (companyname,asof,latest,totalrevenue,cogs,sga,da,netinterest,otherincome," \
                        "taxes,grossprofit,ebit,ebitda,netincome,grossprofitmargin,ebitmargin,ebitdamargin,ebtmargin,netincomemargin) values(" \
                        "'" + sys.argv[4] + "'," + str(int(str(
                    datetime.strptime(data["asof"], "%b%d%Y"))[:4])) + "," + str(
                    latest) + "," + str(AM_IS_I) + "," + str(AM_IS_CORS) + "," + str(AM_IS_EXP) + "," + str(
                    AM_IS_DEP_AMO) + "," + str(AM_IS_NIEXP) + "," + str(AM_IS_OE) + "," + str(AM_IS_TX) + "," + str(
                    gross_profit) + "," + str(ebit) + "," + str(ebitda) + "," + str(netincome) + "," + str(
                    grossprofitmargin) + "," + str(ebitmargin) + "," + str(ebitdamargin) + "," + str(
                    ebitmargin) + "," + str(netincomemargin) + ")"
                cursor.execute(query)
                con.commit()
                break


def map_to_date_obj(date):
    return datetime.strptime(date,"%b%d%Y")
def sort_dict(json_data):
    dates = [date["asof"] for date in json_data["period"]]
    dates = list(map(map_to_date_obj,dates))
    dates.sort()
    return dates

def map_to_string(date):
    return datetime.strftime(date,"%b%d%Y")

dates = sort_dict(json_copy)
dates = list(map(map_to_string,dates))

latest_enum = {}
for i, j in enumerate(dates, -(len(dates) - 1)):
    latest_enum[i] = j
# print(latest_enum,"Latest Enum")
inject_db(json_copy,latest_enum)

